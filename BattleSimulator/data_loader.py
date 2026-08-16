import os
import re
import warnings
import openpyxl

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def resolve_data_dir():
    """
    Dynamically finds the 엑셀 데이터 directory portably.
    1. Checks if Excel files exist in current directory (BattleSimulator)
    2. Checks if Excel files exist in parent directory (2_데이터작업)
    3. Fallback to parent directory or original path
    """
    cur_dir = os.path.dirname(os.path.abspath(__file__))
    target_files = [
        "[전투]레벨디자인 및 몬스터 데이터.xlsx",
        "[전투]포탑 데이터.xlsx",
        "[콘텐츠]트레인 및 승무원 데이터.xlsx"
    ]

    # 1. Check current directory (if BattleSimulator is moved with excel files inside)
    if any(os.path.exists(os.path.join(cur_dir, f)) for f in target_files):
        return cur_dir

    # 2. Check subdirectories (data/, Data/, excel/, Excel/)
    for sub in ["data", "Data", "excel", "Excel"]:
        sub_path = os.path.join(cur_dir, sub)
        if os.path.exists(sub_path) and any(os.path.exists(os.path.join(sub_path, f)) for f in target_files):
            return sub_path

    # 3. Check parent directory (if 2_데이터작업 or upper folder contains excel files)
    parent_dir = os.path.abspath(os.path.join(cur_dir, ".."))
    if any(os.path.exists(os.path.join(parent_dir, f)) for f in target_files):
        return parent_dir

    # 4. Fallback to parent directory if it exists
    if os.path.exists(parent_dir):
        return parent_dir

    return cur_dir

DATA_DIR = resolve_data_dir()

FILES = [
    "[전투]레벨디자인 및 몬스터 데이터.xlsx",
    "[전투]포탑 데이터.xlsx",
    "[콘텐츠]트레인 및 승무원 데이터.xlsx"
]

EXCLUDED_SHEETS = {"factor", "개요", "parlor"}
VALID_TYPES = {"string", "int", "float", "flaot", "vector3", "bool", "double"}

def is_valid_datatype(val):
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    # If contains Korean characters, it's not a standard data type
    if re.search(r'[\uac00-\ud7a3]', s):
        return False
    return s.lower() in VALID_TYPES

def clean_value(val, dtype):
    if val is None:
        return None
    s_val = str(val).strip()
    if s_val.lower() in ("null", "none", ""):
        return None
    
    dtype_clean = dtype.lower()
    try:
        if dtype_clean in ("int", "float", "flaot", "double"):
            f_val = float(val)
            if f_val.is_integer() and dtype_clean == "int":
                return int(f_val)
            return round(f_val, 4)
        elif dtype_clean == "bool":
            return str(val).lower() in ("true", "1", "y", "yes")
        elif dtype_clean == "vector3":
            return str(val)
        else:
            return str(val)
    except Exception:
        return str(val)

class DataLoader:
    def __init__(self, data_dir=DATA_DIR):
        self.data_dir = data_dir
        self.data = {}  # sheet_name -> list of dict records
        self.columns = {}  # sheet_name -> list of (col_name, dtype)
        self.load_all_data()

    def reload_all_data(self):
        self.data.clear()
        self.columns.clear()
        self.load_all_data()

    def load_all_data(self):
        for fname in FILES:
            fpath = os.path.join(self.data_dir, fname)
            if not os.path.exists(fpath):
                print(f"[Warning] File not found: {fpath}")
                continue
            
            try:
                wb = openpyxl.load_workbook(fpath, data_only=True)
                for sname in wb.sheetnames:
                    sname_clean = sname.strip().lower()
                    if sname_clean in EXCLUDED_SHEETS:
                        continue
                    
                    ws = wb[sname]
                    cols, records = self._parse_sheet(ws)
                    self.columns[sname] = cols
                    self.data[sname] = records
                    print(f"[Loaded] Sheet '{sname}': {len(cols)} columns, {len(records)} records")
            except Exception as e:
                print(f"[Error] Failed to load {fname}: {e}")

    def _parse_sheet(self, ws):
        # Row 3: Data type, Row 4: Column name
        max_col = ws.max_column
        row3 = [ws.cell(3, c).value for c in range(1, max_col + 1)]
        row4 = [ws.cell(4, c).value for c in range(1, max_col + 1)]

        valid_cols = []
        for idx, (dtype, colname) in enumerate(zip(row3, row4)):
            if colname is not None and is_valid_datatype(dtype):
                col_type = str(dtype).strip()
                if col_type.lower() == "flaot":
                    col_type = "float"
                valid_cols.append((idx + 1, str(colname).strip(), col_type))

        records = []
        for r in range(5, ws.max_row + 1):
            row_dict = {}
            has_valid_data = False
            for col_idx, colname, dtype in valid_cols:
                raw_val = ws.cell(r, col_idx).value
                cleaned_val = clean_value(raw_val, dtype)
                if cleaned_val is not None:
                    has_valid_data = True
                row_dict[colname] = cleaned_val
            
            # Avoid adding empty padding rows
            if has_valid_data and any(row_dict.values()):
                records.append(row_dict)

        clean_col_info = [(cname, ctype) for _, cname, ctype in valid_cols]
        return clean_col_info, records

    def get_sheet_data(self, sheet_name):
        return self.data.get(sheet_name, [])

    def get_sheet_columns(self, sheet_name):
        return self.columns.get(sheet_name, [])

if __name__ == "__main__":
    loader = DataLoader()
    print("\nSummary of loaded data:")
    for sname, records in loader.data.items():
        print(f" - {sname}: {len(records)} items")
